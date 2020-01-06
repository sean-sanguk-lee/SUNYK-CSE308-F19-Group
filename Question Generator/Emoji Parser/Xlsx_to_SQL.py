# emojis.xlsx contains 96 worksheets (or 96 emoji categories)
# Generate an SQL file depends on the user input of desired category
    # SQL insert query template: INSERT INTO MatchingGame (image, sound, content) values ('...', '...', '...')

import os
import openpyxl as xl


def main():
    infile_dir = os.getcwd() + "/emojis.xlsx"

    ## init workbook
    write_wb = xl.load_workbook(filename=infile_dir)
    sheetnames = ['animal-bird']

    default = get_default_queries(14)
    with open("../SQL to JSON/REDBOOK.sql", 'w') as wf:
        wf.write(default)

        for sheetname in sheetnames:
            ws = write_wb[sheetname]

            first = True
            for line in ws:
                if first:
                    first = False
                    continue

                img, content, sound = line[3], line[1], line[4]
                query_template = 'INSERT INTO MatchingGame (image, sound, content) values '
                query = query_template + "('" + img.value + "','" + sound.value + "','" + content.value + "')"
                wf.write(query + "\n")

    write_wb.save(infile_dir)


def get_default_queries(line_num):
    default, i = '', 0
    path = "../SQL to JSON/REDBOOK.sql"
    for line in open(path, 'r'):
        if i < line_num:
            default += line
            i += 1
    return default


if __name__ == '__main__':
    main()