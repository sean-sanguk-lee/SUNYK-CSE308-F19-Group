import os
import time
import openpyxl as xl

from PIL import Image
from io import BytesIO
import base64
import pyimgur
from gtts import gTTS

CLIENT_ID = '209b7fbc19960da'


def main():
    infile_dir = os.getcwd() + "/emojis.xlsx"

    ## init workbook
    write_wb = xl.load_workbook(filename=infile_dir)
    set_worksheet(write_wb, infile_dir, ['animal-bird'])


def set_worksheet(write_wb, infile_dir, sheetnames):
    for sheetname in sheetnames:
        ws = write_wb[sheetname]

        for line in ws:
            base64_str = line[2].value
            if not base64_str.startswith('Base'):
                # upload base64 to imgur and save the link
                title = line[1].value
                filename = title + '.png'

                uploaded = upload_base64_to_imgur(CLIENT_ID, base64_str, filename)

                ws[line[3].column_letter + str(line[3].row)] = uploaded.link

                # query Google TTS and save .mp3
                tts = gTTS(text=title, lang='en')
                mp3_outdir = './sound/' + title + ".mp3"
                tts.save(mp3_outdir)

                ws[line[4].column_letter + str(line[4].row)] = mp3_outdir

    write_wb.save(infile_dir)


def upload_base64_to_imgur(client_id, base64_str, filename):
    time.sleep(3)
    # im = Image.open(BytesIO(base64.b64decode(base64_str[22:])))
    # im.save(filename, 'PNG')

    PATH = os.getcwd() + '/' + filename
    imgur_client = pyimgur.Imgur(client_id)

    # uploaded = imgur_client.upload_image(PATH, title=filename[:-4])
    uploaded = imgur_client.upload_image(PATH, title=filename)

    # if os.path.exists(filename):
    #     os.remove(filename)

    return uploaded


if __name__ == '__main__':
    main()